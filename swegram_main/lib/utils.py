import codecs
import logging
import os
import shutil
import tempfile
from collections import Counter, OrderedDict, defaultdict
from hashlib import md5
from pathlib import Path
from typing import (
    Any, Callable, Dict, Generator, Iterator, List, Optional,
    TextIO, TypeVar, Tuple, Union
)

from openpyxl import Workbook, worksheet
from openpyxl.reader.excel import load_workbook

from swegram_main.config import XLSX_CONLL_SHEET_PREFIX
from swegram_main.data.metadata import convert_xlsx_labels_to_string, parse_metadata
from swegram_main.lib.converter import Converter
from swegram_main.lib.logger import get_logger


FT = TypeVar("FT", bound=Iterator[Union[Dict[str, str], str]])
FP = TypeVar("FP", bound=Tuple[str, callable, Optional[callable], Dict[str, List[Tuple[str, Any]]], Dict[str, Any]]) # Feature Parameters
K = TypeVar("K", bound=Dict[str, List[Tuple[str, Any]]])
N = TypeVar("N", List[Union[int, float]], Counter)
T = TypeVar("T", bound=List[Tuple[List[List[List[str]]], Dict[str, str]]])
logger = get_logger(__name__)


class MetaFormatError(Exception):
    """Meta format error"""


class ConllFormatError(Exception):
    """conll format error"""


class AnnotationError(Exception):
    """Annotation error"""


class LoadXlsxFormatError(Exception):
    """Load Xlsx Format Error"""


class FeaturePreparationError(Exception):
    """Feature Preparation Error"""


class FileContent:

    def __init__(self, filepath: Path) -> None:
        self.filepath = filepath

    def _convert(self):
        yield from Converter(self.filepath).parse()

    def get(self) -> FT:
        for line in self._convert():
            if not line:
                continue
            metadata = parse_metadata(line)
            if metadata is None:
                yield line
            else:
                yield metadata


class XlsxClient:

    def __init__(self, output_path: Path) -> None:
        self.wb = Workbook()
        self.output_name = output_path

    def dump_cell(self, sheet: worksheet, row: int, column: int, value: str) -> None:
        cell = sheet.cell(row=row, column=column)
        cell.value = value

    def dump_column_list(self, sheet: worksheet, row: int, base_column: int, values: List[str]) -> None:
        for column, value in enumerate(values, base_column):
            self.dump_cell(sheet, row, column, value)


class XlsxAnnotationClient(XlsxClient):

    def dump(self, conll: Path, model: str, language: str, normalized: bool, annotation: str):
        text, labels = read_conll_file(conll)[0]
        meta_sheet = self.wb["Sheet"]
        meta_sheet.title = "Annotation-metadata"
        meta_sheet["A1"] = "Swegram Annotation"
        metadata = OrderedDict({
            "Model": model,
            "Language": language,
            "Normalization": normalized,
            "Annotation": annotation,
        })
        for row, (key, value) in enumerate(metadata.items(), 2):
            self.dump_cell(meta_sheet, row, 1, key)
            self.dump_cell(meta_sheet, row, 2, value)

        # dump conll text in xlsx file
        for index, (text, labels) in enumerate(read_conll_file(conll), 1):
            self.dump_text(text, labels, f"text_{index}")
        self.wb.save(filename=self.output_name)

    def dump_labels(self, sheet: worksheet, labels: Dict[str, str]) -> None:
        self.dump_cell(sheet, 1, 1, "Metadata")
        if labels:
            for index, (key, value) in enumerate(labels.items(), 1):
                self.dump_cell(sheet, 1, index * 2, key)
                self.dump_cell(sheet, 1, index * 2 + 1, value)
        else:
            self.dump_cell(sheet, 1, 2, None)

    def dump_text(self, text: List[List[List[str]]], labels: Dict[str, str], title: str) -> None:
        sheet = self.wb.create_sheet(title=title)
        self.dump_labels(sheet, labels)
        row = 2
        for paragraph in text:
            for sentence in paragraph:
                for token in sentence:
                    fields = token.split("\t")
                    for col, field in enumerate(fields, 1):
                        self.dump_cell(sheet, row, col, field)
                    row += 1
                row += 1
            row += 1

    @staticmethod
    def load_corpus(input_path: Path, outdir_path: Path) -> Path:
        output_path = outdir_path.joinpath(f"{input_path.stem}.conll")
        try:
            wb = load_workbook(input_path)
            with open(output_path, mode="w", encoding="utf-8") as output_file:
                for sheet_name in wb.sheetnames:
                    if sheet_name.startswith(XLSX_CONLL_SHEET_PREFIX):
                        sheet = wb[sheet_name]
                        convert_sheet2conll(sheet, output_file)

        except KeyError as err:
            raise LoadXlsxFormatError(f"Incorrect xlsx conll format: {err}") from err

        return output_path


def convert_sheet2conll(sheet: worksheet, output_file: TextIO) -> None:
    rows = sheet.rows
    label_list = [cell.value for cell in next(rows) if cell.value][1:]
    output_file.write(f"{convert_xlsx_labels_to_string(label_list)}\n")
    for row in rows:
        fields = [c.value for c in row if c.value]
        if fields:
            output_file.write("\t".join(fields))
        else:
            output_file.write("\n")


def get_md5(filepath: Path) -> str:
    """Get md5 value"""
    with codecs.open(filepath, "rb") as inputfile:
        return md5(inputfile.read()).hexdigest()


def get_conll_md5(filepath: Path) -> str:
    with open(filepath, mode="r", encoding="utf-8") as input_file:
        lines = [line for line in input_file.readlines() if not line.startswith("#")]
        return md5("\n".join(lines).encode()).hexdigest()


def get_content_md5(content: str) -> str:
    return md5(content).hexdigest()


def get_size(filepath: Path) -> Optional[int]:
    """Get size"""
    if not os.path.exists(filepath):
        logging.debug(f"{filepath} does not exist.")
        return None
    return os.path.getsize(filepath)


def write(filepath: Path, context: str) -> None:
    """Write into file with context"""
    with codecs.open(filepath, mode="w", encoding="utf-8") as output_file:
        output_file.write(context)


def read(filepath: Path) -> Generator:
    """Read out file given arg into generator"""
    with codecs.open(filepath, mode="r", encoding="utf-8") as input_file:
        line = input_file.readline()
        while line:
            yield line
            line = input_file.readline()


def _initialize_conllu_reading(file_content: FT) -> Tuple[Union[None, Dict[str, str]], Union[str, Dict[str, str]]]:
    component = next(file_content)
    while True:
        # Skip blank lines in the beginning of the file content
        if isinstance(component, str):
            if not component.strip():
                component = next(file_content)
            else:
                return None, component
        elif isinstance(component, dict):
            return component, next(file_content)
        else:
            raise MetaFormatError(f"Invalid format, got {type(component)}:{component}")
    

def read_conll_file(input_path: Path) -> T:
    """Read conllu text"""
    texts: T = []
    paragraphs, sentences, sentence, = [], [], []
    file_content = FileContent(input_path).get()

    def _append_paragraph() -> None:
        if sentence:
            sentences.append(sentence)
        if sentences:
            paragraphs.append(sentences)

    def _append_text(meta: Dict[str, str]) -> None:
        _append_paragraph()
        if paragraphs:
            texts.append((paragraphs, meta))
        elif meta:
            logger.warning(f"Medata data {meta}: text is empty")

    try:  # pylint: disable=too-many-try-statements
        meta, component = _initialize_conllu_reading(file_content)
        while True:
            newline = 0
            while True:
                if isinstance(component, str):
                    if component == "\n":
                        newline += 1
                    elif component.startswith("#"):
                        pass
                    elif not newline:
                        sentence.append(component)
                    elif newline == 2:
                        _append_paragraph()
                        sentences, sentence = [], [component]
                        newline = 0
                    elif newline == 1 and sentence:
                        sentences.append(sentence)
                        sentence = [component]
                        newline = 0
                    else:
                        raise ConllFormatError(f"Too many blank lines, max 2 newlines , but got {newline}")
                elif isinstance(component, dict):
                    _append_text(meta)
                    meta = component
                    paragraphs, sentences, sentence = [], [], []
                    break
                else:
                    raise MetaFormatError(f"Invalid format, got {type(component)}:{component}")
                component = next(file_content)
            component = next(file_content)

    except StopIteration:
        _append_text(meta)
    return texts


def cut(
    func: Callable, filepath: Path, output_path: Optional[Path] = None, append_token_index: bool = False,
    append_text_index: bool = False
) -> None:
    """insertion in the annotation files"""
    lines = read(filepath)
    index = 0
    paragraph_index, sentence_index = 1, 1
    num_newlines = 0
    with tempfile.NamedTemporaryFile(mode="w", encoding="utf-8") as output_file:
        try:  # pylint: disable=too-many-try-statements
            while True:
                line = next(lines)
                if line.strip() and not line.strip().startswith("#"):
                    line = func(line)
                    if append_token_index:
                        if num_newlines > 0:
                            index = 1
                        else:
                            index += 1
                        line = "\t".join([str(index), line])
                            
                    if append_text_index:
                        if num_newlines > 1:
                            paragraph_index += 1
                            sentence_index = 1
                            num_newlines = 0
                        elif num_newlines == 1:
                            sentence_index += 1
                            num_newlines = 0
                        line = "\t".join([f"{paragraph_index}.{sentence_index}", line])
                    output_file.write(line)
                else:
                    output_file.write(line)
                    if line == "\n":
                        num_newlines += 1
        except StopIteration:
            output_file.flush()
            if output_path:
                shutil.copy(output_file.name, output_path)
            else:
                shutil.copy(output_file.name, filepath)


def change_suffix(filepath: Path, suffix: str) -> Path:
    return filepath.parent.joinpath(f"{filepath.stem}{os.path.extsep}{suffix}")


####################################################
#    functions for statistics                      #
####################################################
def mean(numbers: N) -> float:
    if isinstance(numbers, Counter):
        return r2(sum(k * v for k, v in numbers.items()), sum(numbers.values()))
    return r2(sum(numbers), max(len(numbers), 1))


def median(numbers: N) -> float:
    if isinstance(numbers, Counter):
        return median([key for key, value in numbers.items() for _ in range(value)])

    if not numbers: return 0
    numbers.sort()
    if len(numbers) % 2 == 1:
        return numbers[len(numbers) // 2]
    index = len(numbers) // 2
    a, b = numbers[index-1: index+1]
    return r2((a + b) / 2)


def r2(number: int, *args) -> float:
    """keep two decimals"""
    if not args or not args[0]:
        return round(number, 2)
    number2, *_ = args
    return round(number / number2, 2)


def merge_counters(blocks: List[object], field: str) -> Counter:
    counter_instance = Counter()
    for block in blocks:
        counter_instance.update(getattr(block, field))
    return counter_instance


def merge_counters_for_fields(blocks: List[object], fields: List[str]) -> List[Counter]:
    return [merge_counters(blocks, field) for field in fields]


def merge_dicts(blocks: List[object], field: str, data_type: type = int) -> defaultdict:
    df = defaultdict(data_type)
    for block in blocks:
        for key, value in getattr(block, field).items():
            if issubclass(data_type, int):
                df[key] += value
            elif issubclass(data_type, list):
                df[key].extend(value)
            elif issubclass(data_type, set):
                df[key] = df[key].union(value)
            elif issubclass(data_type, dict):
                df[key].update(value)
            else:
                raise AnnotationError(f"Unsopported data type: {data_type}")
    return df


def merge_dicts_for_fields(blocks: List[object], fields: List[str]) -> List[defaultdict]:
    defaultdicts = []
    for field in fields:
        defaultdicts.append(merge_dicts(blocks, field))
    return defaultdicts


def merge_digits(blocks: List[object], field: str, operation: callable = sum) -> int:
    return operation([getattr(block, field) for block in blocks])


def merge_digits_for_fields(blocks: List[object], fields: List[str], operation: callable = sum) -> int:
    return [merge_digits(blocks, field, operation) for field in fields]


def mixin_merge_digits_or_counters(blocks: List[object], field: str, operation: callable = sum) -> Any:
    if isinstance(getattr(blocks[0], field), Counter):
        return merge_counters(blocks, field)
    return merge_digits(blocks, field, operation)


def mixin_merge_digits_or_dicts(
    blocks: List[object], field: str,
    operation: callable = sum, data_type: type = int
) -> Any:
    if isinstance(getattr(blocks[0], field), defaultdict):
        return merge_dicts(blocks, field, data_type) 
    return merge_digits(blocks, field, operation)


def get_path(index: int, heads: List[Union[int, None]]) -> List[int]:
    path = [index]
    while index:
        path.append(heads[index])
        if len(path) > len(heads):
            raise AnnotationError(f"Failed to get dependency path for index {index} in heads {heads}")
        index = heads[index]
    return path


def get_child_nodes(index: int, heads: List[Union[int, None]]) -> List[int]:
    """Get all direct and indirect child nodes given index node"""
    child_nodes = []
    while index in heads:
        child = heads.index(index)
        child_nodes.append(child)
        heads[child] = None
        for c in get_child_nodes(child, heads):
            child_nodes.append(c)
    return child_nodes


def is_a_ud_tree(heads: List[str], error_prefix: str = "") -> Union[bool, str]:
    heads = [int(head) for head in heads]
    children = list(range(1, len(heads)+1))
    if 0 not in heads:
        return f"{error_prefix} Root is missing."
    if Counter(heads)[0] > 1:
        return f"{error_prefix} More than one roots in sentence"
    if len(children) > len(set(children)):
        return f"{error_prefix} Same indeces for two nodes in sentence."

    for i in children:
        head = [heads[i - 1]]
        while 0 not in head:
            if heads[head[-1] - 1] in head:
                return f"{error_prefix} Cycle Error in sentence."
            head.append(heads[head[-1] - 1])
        head = []
    return True


def prepare_feature(*args: Union[str, callable]) -> FP:
    """Two arg types are defined in this case

    with arg as arg_type, it can be used in the incsc computation;
    with attribute as arg_type, it needs to be converted from the instance before it is utilized for computation
    with attribute_arg as arg_type, the argument used when converting the attribute into arg for feature computation
    """
    feature_name, feature_func, attribute_func, *rest_args = args
    kwargs, attribute_kwargs = {}, {}
    while rest_args:
        try:
            arg_key, arg_value, arg_type, *rest_args = rest_args
            if arg_type == "attribute_arg":
                attribute_kwargs.update({arg_key: arg_value})
            elif arg_type in kwargs:
                kwargs[arg_type].append((arg_key, arg_value))
            else:
                kwargs[arg_type] = [(arg_key, arg_value)]
        except ValueError as err:
            raise FeaturePreparationError(f"Failed to parse args for feature, args: {args}") from err
    return feature_name, feature_func, attribute_func, kwargs, attribute_kwargs


def parse_args(kwarg_list: K, func: callable, content: Any, **kwargs) -> Dict[str, Any]:
    args = kwarg_list.get("arg", [])
    args.extend([
        (
            key,
            func(content, attribute, **kwargs)
        ) for key, attribute in kwarg_list.get("attribute", [])
    ])
    return dict(args)


def incsc(c: Union[int, float], t: Union[int, float]) -> float:
    try:
        return r2(c * 1000, t)
    except ZeroDivisionError:
        return 1000.00
